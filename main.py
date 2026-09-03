import re
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

# ========================================
# 設定
# ========================================

INPUT_FILE = Path("input/teams_yyyymmdd.txt")
EXCEL_FILE = Path("data/WorkList.xlsx")

# ========================================
# 入力ファイル読込
# ========================================

with open(INPUT_FILE, encoding="utf-8") as f:
    lines = [line.strip() for line in f if line.strip()]

if len(lines) < 3:
    raise ValueError("入力データが不足しています")

# ========================================
# 日付取得
# ========================================

header = lines[0]

m = re.search(
    r"(\d{4})年(\d{2})月(\d{2})日",
    header
)

if not m:
    raise ValueError("日付が取得できません")

sheet_name = f"{m.group(1)}{m.group(2)}{m.group(3)}"
sheet_date = datetime.strptime(sheet_name, "%Y%m%d")

print(f"処理日: {sheet_name}")

# ========================================
# Excel作成／読込
# ========================================

if EXCEL_FILE.exists():
    wb = load_workbook(EXCEL_FILE)
else:
    wb = Workbook()

    # デフォルトシート名変更
    ws = wb.active
    ws.title = "input"

# ========================================
# inputシート作成
# ========================================

if "input" in wb.sheetnames:
    ws_input = wb["input"]

    max_row = ws_input.max_row

    if max_row > 0:
        ws_input.delete_rows(1, max_row)
else:
    ws_input = wb.create_sheet("input")

# inputシートへ全データ出力
for row_no, text in enumerate(lines, start=1):
    ws_input.cell(row=row_no, column=1, value=text)

# ========================================
# 日付シート作成 indexは0始まりなので、左から2番目なら1
# ========================================

if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    max_row = ws.max_row
    if max_row > 0:
        ws.delete_rows(1, max_row)
else:
    ws = wb.create_sheet(sheet_name, 1)

# ========================================
# ヘッダー
# ========================================

headers = [
    ("日付", 10),
    ("拠点", 12),
    ("環境名", 30),
    ("開始日", 10),
    ("連続稼働日数", 12)
]

header_font = Font(bold=True)

header_fill = PatternFill(
    fill_type="solid",
    fgColor="D9D9D9"
)

header_alignment = Alignment(horizontal="center")

for col_no, (header_text, width) in enumerate(headers, start=1):

    cell = ws.cell(
        row=1,
        column=col_no,
        value=header_text
    )

    # ヘッダー書式
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

    # 列幅設定
    column_letter = cell.column_letter
    ws.column_dimensions[column_letter].width = width

# ヘッダー固定
ws.freeze_panes = "A2"
        
# ========================================
# 前日シート取得
# ========================================

today = datetime.strptime(sheet_name, "%Y%m%d")
yesterday = today - timedelta(days=1)

prev_sheet_name = yesterday.strftime("%Y%m%d")

prev_data = {}

if prev_sheet_name in wb.sheetnames:

    prev_ws = wb[prev_sheet_name]

    for row in range(2, prev_ws.max_row + 1):

        env_name = prev_ws.cell(row, 3).value

        start_date = prev_ws.cell(row, 4).value

        consecutive_days = prev_ws.cell(row, 5).value

        prev_data[env_name] = (
            start_date,
            consecutive_days
        )

    print(f"前日シート発見: {prev_sheet_name}")

else:

    print(f"前日シートなし: {prev_sheet_name}")

# ========================================
# データ出力
# ========================================

excel_row = 2

for i in range(1, len(lines), 2):

    if i + 1 >= len(lines):
        break

    location = lines[i]
    env_name = lines[i + 1]

    date_cell = ws.cell(excel_row, 1)
    date_cell.value = sheet_date
    date_cell.number_format = "yyyymmdd"
    ws.cell(excel_row, 2, location)
    ws.cell(excel_row, 3, env_name)

    # 開始日 連続稼働日数

    if env_name in prev_data:

        start_date, consecutive_days = prev_data[env_name]

        # ws.cell(excel_row, 4, start_date)
        start_cell = ws.cell(excel_row, 4)
        if isinstance(start_date, datetime):
        #    print(f"開始日ｉｆ : {start_date}")
            start_cell.value = start_date
        else:
            print(f"開始日ｅｌｓｅ : {start_date}")
            start_cell.value = datetime.strptime(
                str(start_date),
                "%Y%m%d"
            )
        start_cell.number_format = "yyyymmdd"
        
        ws.cell(excel_row, 5, consecutive_days + 1)

    else:

        start_cell = ws.cell(excel_row, 4)
        start_cell.value = sheet_date
        start_cell.number_format = "yyyymmdd"
        ws.cell(excel_row, 5, 1)

    excel_row += 1

# ========================================
# 保存
# ========================================

alert_count = 0

for row in range(2, ws.max_row + 1):

    days = ws.cell(row, 5).value

    if days >= 3:
        alert_count += 1

wb.save(EXCEL_FILE)

print("WorkList.xlsx 更新完了")
print(f"シート名 : {sheet_name}")
print(f"件数     : {excel_row - 2}")
print(f"3日以上連続稼働: {alert_count}件")
